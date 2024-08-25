"""Utilities for manipulating Multiple Sequence Alignment
    and comparing the characters at each position to account for
    mutations.
    """

import os
import sys
from collections import OrderedDict

from Bio import SeqIO
from openpyxl.workbook.child import _WorkbookChild


class MultiFastaMutationsFinder:
    """Creates blueprint for handling Multi Fasta Alignment files and extracting
    mutations is any from comparison with a reference sequence and writing them to
    an Excel sheet.
    """

    def __init__(
        self,
        path: str,
        sheet: _WorkbookChild,
        protein_names: list,
        extension: str,
    ) -> None:
        """Constructor

        Args:
            path (str): Path to alignment data
            sheet (_Workbook child): An OpenpyXL Workbook child
            protein_names (list): A list of the protein names extracted from the file basename
            extension (str): The extension of the protein name files
        """
        self.path = path
        self.sheet = sheet
        self.ref_ids = ["H37Rv", "CDC1551", "F11", "H37Ra", "Erdman", "HN878", "KZN 1435"]
        self.protein_names = protein_names
        self.extension = extension
        self.id_mutations = OrderedDict()
        self.existing_ids = {}

    def process_fasta_file(self) -> None:
        """Process the Multi Fasta Alignment file. Parse it using BioPython's
        SeqIO.parse() method and then compare sequences to that of the reference.
        """
        for protein in self.protein_names:
            file = os.path.join(self.path, protein + self.extension)
            with open(file, "r", encoding="utf-8") as handle:
                id_sequences_dict = OrderedDict()
                for record in SeqIO.parse(handle, "fasta"):
                    id_sequences_dict[record.id] = record.seq

            print(f"Comparing Aligned Sequences for mutations in {protein}")

            # Retrieve the reference sequence
            reference_id = self.find_reference(id_sequences_dict)
            ref_seq = id_sequences_dict.get(reference_id, "")

            # Check if the reference sequence exists
            if not ref_seq:
                print("ERROR: Reference ID and sequence not found!\n"
                      "Check your .mfa file for reference ID and sequence.\n"
                      f"If not in this list {self.ref_ids}, then kindly enter it below.\n")

                while not ref_seq:
                    reference_id = input("Reference ID: ").strip()
                    if reference_id.lower() == "exit":
                        print("Exiting...")
                        sys.exit()

                    ref_seq = id_sequences_dict.get(reference_id, "")

                    if not ref_seq:
                        print("Invalid Reference ID. Please try again or type 'exit' to quit.")
            else:
                for record_id, record_seq in id_sequences_dict.items():
                    self.handle_protein_ids(protein, record_id, record_seq, ref_seq)

        print("Done!")

    def handle_protein_ids(self, protein, record_id, record_seq, ref_seq) -> None:
        """
        Check if protein exits in mutation dictionary, if yes, proceed to
        compare aligned sequences for mutation else set protein as the key with
        a list of mutation(s) as values, if they do exist.

        Args:
            protein (str): The protein name
            record_id (str): The protein ID
            record_seq (str): The protein sequence
            ref_seq (str): The reference sequence
        @return: None
        """
        # Handle existing ID
        if (protein in self.id_mutations) and (
                record_id not in self.ref_ids
        ):
            self.compare_aligned_sequences(
                protein, record_id, record_seq, ref_seq
            )
        elif record_id not in self.ref_ids:
            self.id_mutations[protein] = []
            self.compare_aligned_sequences(
                protein, record_id, record_seq, ref_seq
            )

    def compare_aligned_sequences(
        self, protein: str, record_id: str, isolate_seq: str, ref_seq: str
    ) -> None:
        """Compare aligned sequences to a reference sequence.
        If the isolate's sequence is an "X" or matches that in the reference sequence, then write "X", else, write
        the mutation that occurs, like so, original nucleotide in reference first, then
        position where mutation occurs, then mutated nucleotide in isolate.
        Mutations are stored in a dictionary as a tuple with their ID, in a list, whose
        key is the protein for which they code for.

        Args:
            protein (str): The protein in question.
            Gotten from file name
            record_id (str): The ID of the isolate sequence
            isolate_seq (str): The nucleotides that make the protein (gene) in question
            ref_seq (str): The nucleotides that make the protein in question but in the reference genome
        """
        mutations_list = []
        for i, _ in enumerate(isolate_seq):
            if isolate_seq[i] == "X" or (
                isolate_seq[i] == ref_seq[i]
            ):  # Check if ID was missing in previous files
                mutations_list.append("X")
            elif isolate_seq[i] != ref_seq[i]:
                mutations_list.append(f"{ref_seq[i]}{i+1}{isolate_seq[i]}")

        # Check if the mutation list is not full of X's, then join them by ";"
        mutations = ";".join(filter(lambda x: x != "X", mutations_list))
        if mutations:
            self.id_mutations[protein].append((record_id, mutations))  # Add mutations
        else:
            self.id_mutations[protein].append((record_id, "X"))

    def get_mutations(self):
        """Prints out all mutations.
        A dictionary with protein as a key and a list
        of tuples which consists of Isolate ID and mutation.
        """
        print(f"{self.id_mutations}")

    def insert_to_excel(self) -> None:
        """Insert mutations to excel corresponding to their IDs. This is done
        by iterating through the mutation list, then finding the Isolate's ID
        from a dictionary of existing IDs and then extracting the
        value (position of ID in an Excel sheet), after which one simply inserts
        mutation at position of the ID.
        """
        for column_index, mutation_list in enumerate(self.id_mutations.values()):
            for mutation in mutation_list:
                # Find ID's position in sheet and insert mutation
                position_in_sheet = self.existing_ids[mutation[0]]
                # insert into excel
                self.sheet.cell(
                    row=position_in_sheet, column=column_index + 2
                ).value = mutation[1]

        # Not all IDs will belong to each protein alignment file.
        # Hence, some cells will be empty for those. In such cases,
        # fill them with an "X"
        last_column = self.sheet.max_column
        for row in self.sheet.iter_rows(min_row=2):
            for col_index in range(last_column):
                cell = row[col_index]
                if cell.value is None:
                    cell.value = "X"

        print("Done!")

    def insert_ids_to_excel(self) -> None:
        """Insert headers and IDs of Isolates into an Excel sheet.
        Store existing IDs that are in the Excel sheet in a dictionary, with IDs as keys and their
        position in the sheet as value.
        Helps when inserting mutations in a sheet.
        """
        headers = self.protein_names.copy()
        headers.insert(0, "Isolate ID")
        for header_index, header_data in enumerate(headers):
            self.sheet.cell(row=1, column=1 + header_index).value = header_data

        for mutation_list in self.id_mutations.values():
            for row_index, mutation_tuple in enumerate(mutation_list):
                # Check if Isolate ID is not a part of the existing Isolate IDs,
                # then insert mutation to cell. Then insert ID to existing ID
                # dictionary for later reference.
                if mutation_tuple[0] not in self.existing_ids:
                    new_row = row_index + 2
                    self.sheet.insert_rows(new_row)
                    self.sheet.cell(row=new_row, column=1).value = mutation_tuple[0]
                    self.existing_ids[mutation_tuple[0]] = new_row

            # After each insertion, inconsistencies in ID positions in the existing ID
            # dictionary.
            # So iterate through rows in the sheet and do it again.
            for row_index, row in enumerate(
                self.sheet.iter_rows(min_row=2, values_only=True)
            ):
                if row and row[0]:
                    self.existing_ids[row[0]] = row_index + 2

        print("Done!")

    def find_reference(self, sequence_dict) -> str | None:
        """
        Check reference IDs in ref_ids dictionary if any exist in the
        sequence dictionary, then return that ID.

        Args:
            sequence_dict (dict): A dictionary with nucleotide IDs as keys and
            sequences as values.
        @return: str | None
        """
        for ref_id in self.ref_ids:
            if ref_id in sequence_dict:
                return ref_id
        return None