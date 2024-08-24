"""Utilities for creating, loading, writing, saving, finding 
    occupied rows and columns and inserting headers to Spreadsheets
    """

import os
import pandas as pd
from openpyxl import Workbook, load_workbook


def create_workbook():
    """Function creates a workbook from which a
    Spreadsheet is made active for manipulation

        Returns:
            tuple: a Workbook object and Workbook child (sheet)
    """
    workbook = Workbook()
    sheet = workbook.active
    return workbook, sheet


def save_worksheet(workbook: Workbook, filename: str):
    """Save changes to a worksheet

    Args:
        workbook (Workbook): Workbook object to invoke save method
        filename (str): Name to save the output file as.
        Specify a path to save at a particular place
    """
    workbook.save(filename)


def extract_gene_name_from_file(path: str, extension: str = ".mfa"):
    """Extract the name of the proteins from their multiple alignment file (MAF)
    filenames, which are saved by the Gene name and store them in a list

    Args:
        path (str): The directory where the protein files reside
        extension (str, optional): The extension of the multiple alignment files.
                                Default to '.mfa'
    """
    filenames = []
    for filename in os.listdir(path):
        if filename.endswith(extension):
            base_filename = os.path.splitext(filename)[0]
            filenames.append(base_filename)

    return filenames


def insert_headers_to_excel(header, sheet, start_col=1):
    for header_index, header_data in enumerate(header):
        sheet.cell(row=1, column=start_col + header_index).value = header_data


def insert_data_to_excel(mutations: str, sheet, start_row: int, start_col: int):
    """Insert compared sequences into the sheet.
    The compared sequences will normally arrive as a list of tuples
    from the compared_aligned_sequences function, which returns the
    sequence ID and the compared sequences.

    Args:
        mutations (list): A list of tuples of compared sequences.
        sheet (Workbook child): The sheet in question.
        start_row (int, optional): The row from which to start writing to.
        Default to 1.
        start_col (int, optional): The column to start writing to.
        Default to 1.
    """
    # Write mutation to sheet cell.
    sheet.cell(row=2 + start_row, column=2 + start_col).value = mutations


def insert_data_to_specific_row():
    # Load the workbook
    wb = load_workbook("hello.xlsx")

    # Get the worksheet you want to modify
    worksheet = wb["alpha"]  # Replace "Sheet1" with your actual sheet name

    # Define the position where you want to insert the row (e.g., row 5)
    insert_row = 5

    # Insert the new row (this creates empty cells)
    worksheet.insert_rows(insert_row)

    # Get the data you want to insert in the new row (replace with your actual data)
    new_data = ["Middle1", "Middle2"]

    # Loop through the new data and assign it to the corresponding cells in the inserted row
    for col_idx, value in enumerate(new_data):
        worksheet.cell(row=insert_row, column=col_idx + 1).value = (
            value  # Adjust column number as needed
        )

def excel_to_csv(excel_file, sheet_name):
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    base_filename = os.path.splitext(excel_file)[0]
    csv_file = base_filename + '.csv'
    df.to_csv(csv_file, index=False)
